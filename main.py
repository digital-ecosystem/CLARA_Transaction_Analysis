"""
CLARA Transaction Analysis System - REST API

FastAPI-basierte REST API für Transaktionsanalyse
"""

import uvicorn
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from datetime import datetime
from typing import List, Optional
import pandas as pd
import io
import time
import logging
import os
from pathlib import Path

from models import (
    Transaction, CustomerRiskProfile, AnalysisResponse,
    HealthResponse, RiskLevel
)
from analyzer import TransactionAnalyzer

# Logging Setup
log_dir = Path("logs")
log_dir.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_dir / f'clara_{datetime.now().strftime("%Y%m%d")}.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("CLARA")

# Excel-Unterstützung
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl nicht verfügbar - Excel-Export wird nicht funktionieren")


# Initialisiere FastAPI App
app = FastAPI(
    title="CLARA Transaction Analysis System",
    description="Anti-Smurfing und Transaktionsanalyse mit Weight-Variable, Shannon-Entropie und Trust Score",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# CORS Middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Globaler Analyzer (in Produktion: mit Datenbank-Persistenz)
analyzer = TransactionAnalyzer()

# Start-Zeit für Uptime
start_time = time.time()

# Output-Verzeichnis für generierte CSVs
output_dir = Path("output")
output_dir.mkdir(exist_ok=True)


@app.get("/", response_class=HTMLResponse)
async def root():
    """Root Endpoint - Serves the Web UI"""
    html_path = Path("templates/index.html")
    if html_path.exists():
        with open(html_path, 'r', encoding='utf-8') as f:
            return HTMLResponse(content=f.read())
    else:
        return HTMLResponse(content="<h1>CLARA System</h1><p>UI not found. Please check templates/index.html</p>")


@app.get("/health", response_model=HealthResponse)
async def health_check():
    """Health Check Endpoint"""
    return HealthResponse(
        status="healthy",
        version="1.0.0",
        uptime_seconds=time.time() - start_time
    )


@app.post("/api/analyze/transaction", response_model=CustomerRiskProfile)
async def analyze_single_transaction(transaction: Transaction):
    """
    Analysiert eine einzelne Transaktion
    
    Fügt die Transaktion zum System hinzu und gibt das Risikoprofil zurück.
    """
    try:
        # Füge Transaktion hinzu
        analyzer.add_transactions([transaction])
        
        # Analysiere Kunde
        profile = analyzer.analyze_customer(transaction.customer_id)
        
        return profile
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/analyze/csv", response_model=AnalysisResponse)
async def analyze_csv_file(
    file: UploadFile = File(...),
    recent_days: int = Form(30),
    historical_days: int = Form(365)
):
    """
    Analysiert Transaktionen aus CSV-Datei
    
    Unterstützt zwei CSV-Formate:
    1. Englisch: customer_id,transaction_id,customer_name,transaction_amount,payment_method,transaction_type,timestamp
    2. Deutsch: Datum,Uhrzeit,Timestamp,Kundennummer,Unique Transaktion ID,Vollständiger Name,Auftragsvolumen,In/Out,Art
    
    Parameter:
    - file: CSV-Datei mit Transaktionen
    - recent_days: Anzahl Tage für "recent" Analyse (default: 30, für historische Daten empfohlen: 3650)
    - historical_days: Anzahl Tage für historische Baseline (default: 365)
    
    Zeitstempel ist optional (Format: YYYY-MM-DD HH:MM:SS oder ISO 8601 bzw. DD.MM.YYYY)
    """
    try:
        # Lese CSV (mit Encoding-Detection für Windows CSV-Dateien)
        contents = await file.read()
        
        # Versuche verschiedene Encodings
        df = None
        encodings = ['utf-8', 'windows-1252', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                decoded_content = contents.decode(encoding)
                df = pd.read_csv(io.StringIO(decoded_content))
                break  # Erfolgreich - verwende dieses Encoding
            except (UnicodeDecodeError, Exception) as e:
                continue  # Versuche nächstes Encoding
        
        if df is None:
            raise HTTPException(status_code=400, detail="CSV konnte mit keinem bekannten Encoding gelesen werden")
        
        # Prüfe welches Format vorliegt
        is_new_format = 'Kundennummer' in df.columns
        
        transactions = []
        
        if is_new_format:
            # NEUES FORMAT: Deutsche Spalten
            print(f"[INFO] Verarbeite NEUES Format (deutsch) mit {len(df)} Zeilen")
            
            for idx, row in df.iterrows():
                try:
                    # Mapping: Deutsch → Englisch
                    customer_id = str(row['Kundennummer'])
                    transaction_id = str(row['Unique Transaktion ID'])
                    customer_name = str(row['Vollständiger Name'])
                    
                    # Betrag: Komma → Punkt konvertieren
                    amount_str = str(row['Auftragsvolumen']).replace(',', '.')
                    transaction_amount = float(amount_str)
                    
                    # Zahlungsmethode: "Kredit" → "Kreditkarte"
                    art = str(row['Art']).strip()
                    if art == "Kredit":
                        payment_method = "Kreditkarte"
                    elif art == "Bar":
                        payment_method = "Bar"
                    elif art == "SEPA":
                        payment_method = "SEPA"
                    else:
                        payment_method = art  # Fallback
                    
                    # Transaktionstyp: "In/Out" → "investment/auszahlung"
                    in_out = str(row['In/Out']).strip()
                    if in_out == "In":
                        transaction_type = "investment"
                    elif in_out == "Out":
                        transaction_type = "auszahlung"
                    else:
                        transaction_type = "investment"  # Fallback
                    
                    # Timestamp: Kombiniere Datum + Uhrzeit oder nutze Timestamp-Spalte
                    timestamp = None
                    if 'Timestamp' in df.columns and pd.notna(row['Timestamp']):
                        try:
                            # Format: DD.MM.YYYY
                            date_str = str(row['Timestamp'])
                            timestamp = pd.to_datetime(date_str, format='%d.%m.%Y')
                            
                            # Füge Uhrzeit hinzu falls vorhanden
                            if 'Uhrzeit' in df.columns and pd.notna(row['Uhrzeit']):
                                try:
                                    # Uhrzeit ist ein Dezimalwert (0.663... = Stundenanteil des Tages)
                                    time_str = str(row['Uhrzeit']).replace(',', '.')
                                    time_decimal = float(time_str)
                                    hours = int(time_decimal * 24)
                                    minutes = int((time_decimal * 24 - hours) * 60)
                                    seconds = int(((time_decimal * 24 - hours) * 60 - minutes) * 60)
                                    timestamp = timestamp.replace(hour=hours, minute=minutes, second=seconds)
                                except Exception as e:
                                    pass  # Behalte nur Datum
                        except Exception as e:
                            print(f"[WARN] Timestamp-Parse-Fehler in Zeile {idx+2}: {e}")
                            timestamp = None
                    
                    txn = Transaction(
                        customer_id=customer_id,
                        transaction_id=transaction_id,
                        customer_name=customer_name,
                        transaction_amount=transaction_amount,
                        payment_method=payment_method,
                        transaction_type=transaction_type,
                        timestamp=timestamp
                    )
                    transactions.append(txn)
                except Exception as e:
                    print(f"[WARN] Fehler beim Parsen von Zeile {idx+2}: {e}")
                    continue
        
        else:
            # ALTES FORMAT: Englische Spalten
            print(f"[INFO] Verarbeite ALTES Format (englisch) mit {len(df)} Zeilen")
            
            required_columns = [
                'customer_id', 'transaction_id', 'customer_name',
                'transaction_amount', 'payment_method', 'transaction_type'
            ]
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise HTTPException(
                    status_code=400,
                    detail=f"Fehlende Spalten: {', '.join(missing_columns)}"
                )
            
            for idx, row in df.iterrows():
                try:
                    timestamp = None
                    if 'timestamp' in df.columns and pd.notna(row['timestamp']):
                        try:
                            timestamp = pd.to_datetime(row['timestamp'])
                        except:
                            timestamp = None
                    
                    txn = Transaction(
                        customer_id=str(row['customer_id']),
                        transaction_id=str(row['transaction_id']),
                        customer_name=str(row['customer_name']),
                        transaction_amount=float(row['transaction_amount']),
                        payment_method=str(row['payment_method']),
                        transaction_type=str(row['transaction_type']),
                        timestamp=timestamp
                    )
                    transactions.append(txn)
                except Exception as e:
                    print(f"[WARN] Fehler beim Parsen von Zeile {idx+2}: {e}")
                    continue
        
        if not transactions:
            raise HTTPException(
                status_code=400,
                detail="Keine gültigen Transaktionen in CSV gefunden"
            )
        
        print(f"[OK] {len(transactions)} Transaktionen erfolgreich geparst")
        print(f"[INFO] Analyse-Parameter: recent_days={recent_days}, historical_days={historical_days}")
        
        # Erstelle neuen Analyzer mit den spezifischen Parametern
        custom_analyzer = TransactionAnalyzer(
            alpha=0.6,
            beta=0.4,
            historical_days=historical_days
        )
        
        # Füge Transaktionen hinzu
        custom_analyzer.add_transactions(transactions)
        
        # Debug: Prüfe wie viele Kunden wir haben
        num_customers = len(custom_analyzer.transaction_history)
        print(f"[DEBUG] {num_customers} Kunden in transaction_history gefunden")
        
        # Debug: Zeige erste paar Kunden mit Transaktionszahlen
        for i, (cust_id, txns) in enumerate(list(custom_analyzer.transaction_history.items())[:5]):
            print(f"  - Kunde {cust_id}: {len(txns)} Transaktionen")
        
        # Analysiere alle Kunden mit dem spezifizierten Zeitfenster
        profiles = custom_analyzer.analyze_all_customers(recent_days=recent_days)
        
        # Filtere flagged customers (YELLOW, ORANGE, RED)
        flagged = [
            p for p in profiles
            if p.risk_level != RiskLevel.GREEN
        ]
        
        # Zähle nach Risk Level
        summary = {
            "green": sum(1 for p in profiles if p.risk_level == RiskLevel.GREEN),
            "yellow": sum(1 for p in profiles if p.risk_level == RiskLevel.YELLOW),
            "orange": sum(1 for p in profiles if p.risk_level == RiskLevel.ORANGE),
            "red": sum(1 for p in profiles if p.risk_level == RiskLevel.RED),
        }
        
        return AnalysisResponse(
            status="success",
            message=f"{len(transactions)} Transaktionen analysiert, {len(profiles)} Kunden bewertet",
            analyzed_customers=len(profiles),
            flagged_customers=flagged,
            summary=summary
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fehler bei CSV-Verarbeitung: {str(e)}")


@app.get("/api/customer/{customer_id}/risk-profile", response_model=CustomerRiskProfile)
async def get_customer_risk_profile(
    customer_id: str,
    recent_days: int = 30
):
    """
    Holt das Risikoprofil eines spezifischen Kunden
    
    Args:
        customer_id: Kunden-ID
        recent_days: Zeitfenster für Analyse (Standard: 30 Tage)
    """
    try:
        profile = analyzer.analyze_customer(customer_id, recent_days=recent_days)
        return profile
    
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/flagged-customers", response_model=List[CustomerRiskProfile])
async def get_flagged_customers(
    min_risk_level: Optional[str] = "YELLOW",
    limit: Optional[int] = 100
):
    """
    Holt alle auffälligen Kunden
    
    Args:
        min_risk_level: Minimales Risiko-Level (GREEN, YELLOW, ORANGE, RED)
        limit: Maximale Anzahl Ergebnisse
    """
    try:
        # Analysiere alle Kunden
        profiles = analyzer.analyze_all_customers()
        
        # Filtere nach Risk Level
        risk_levels = ["GREEN", "YELLOW", "ORANGE", "RED"]
        min_index = risk_levels.index(min_risk_level.upper())
        
        flagged = [
            p for p in profiles
            if risk_levels.index(p.risk_level.value) >= min_index
        ]
        
        # Limitiere
        flagged = flagged[:limit]
        
        return flagged
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/statistics")
async def get_statistics():
    """
    Holt System-Statistiken
    """
    try:
        total_customers = len(analyzer.transaction_history)
        total_transactions = sum(
            len(txns) for txns in analyzer.transaction_history.values()
        )
        
        # Analysiere alle
        profiles = analyzer.analyze_all_customers()
        
        risk_distribution = {
            "green": sum(1 for p in profiles if p.risk_level == RiskLevel.GREEN),
            "yellow": sum(1 for p in profiles if p.risk_level == RiskLevel.YELLOW),
            "orange": sum(1 for p in profiles if p.risk_level == RiskLevel.ORANGE),
            "red": sum(1 for p in profiles if p.risk_level == RiskLevel.RED),
        }
        
        avg_suspicion = sum(p.suspicion_score for p in profiles) / len(profiles) if profiles else 0
        
        return {
            "total_customers": total_customers,
            "total_transactions": total_transactions,
            "risk_distribution": risk_distribution,
            "average_suspicion_score": round(avg_suspicion, 2),
            "flagged_percentage": round(
                (risk_distribution["yellow"] + risk_distribution["orange"] + risk_distribution["red"]) / 
                max(total_customers, 1) * 100, 2
            )
        }
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/reset")
async def reset_system():
    """
    Setzt das System zurück (löscht alle Daten)
    
    ⚠️ ACHTUNG: Nur für Testing/Development
    """
    global analyzer
    analyzer = TransactionAnalyzer()
    
    return {
        "status": "success",
        "message": "System wurde zurückgesetzt"
    }


@app.post("/api/batch/transactions", response_model=AnalysisResponse)
async def analyze_transaction_batch(transactions: List[Transaction]):
    """
    Analysiert einen Batch von Transaktionen (als JSON)
    
    Alternative zu CSV-Upload - nimmt JSON-Array von Transaktionen
    """
    try:
        if not transactions:
            raise HTTPException(
                status_code=400,
                detail="Keine Transaktionen bereitgestellt"
            )
        
        # Füge Transaktionen hinzu
        analyzer.add_transactions(transactions)
        
        # Analysiere alle Kunden
        profiles = analyzer.analyze_all_customers()
        
        # Filtere flagged customers
        flagged = [
            p for p in profiles
            if p.risk_level != RiskLevel.GREEN
        ]
        
        # Zähle nach Risk Level
        summary = {
            "green": sum(1 for p in profiles if p.risk_level == RiskLevel.GREEN),
            "yellow": sum(1 for p in profiles if p.risk_level == RiskLevel.YELLOW),
            "orange": sum(1 for p in profiles if p.risk_level == RiskLevel.ORANGE),
            "red": sum(1 for p in profiles if p.risk_level == RiskLevel.RED),
        }
        
        return AnalysisResponse(
            status="success",
            message=f"{len(transactions)} Transaktionen analysiert",
            analyzed_customers=len(profiles),
            flagged_customers=flagged,
            summary=summary
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/analyze/csv-upload")
async def analyze_csv_upload(file: UploadFile = File(...)):
    """
    Analysiert CSV und erstellt automatisch eine markierte CSV-Datei
    
    Returns:
        JSON mit Analyse-Zusammenfassung und Dateinamen der generierten CSV
    """
    try:
        logger.info(f"CSV-Upload gestartet: {file.filename}")
        
        # Lese CSV
        contents = await file.read()
        
        # Versuche verschiedene Encodings
        df = None
        encodings = ['utf-8', 'windows-1252', 'latin-1', 'cp1252']
        
        for encoding in encodings:
            try:
                decoded_content = contents.decode(encoding)
                df = pd.read_csv(io.StringIO(decoded_content))
                logger.info(f"CSV erfolgreich mit {encoding} gelesen")
                break
            except (UnicodeDecodeError, Exception):
                continue
        
        if df is None:
            raise HTTPException(status_code=400, detail="CSV konnte nicht gelesen werden")
        
        # Parse Transaktionen (neues deutsches Format)
        is_new_format = 'Kundennummer' in df.columns
        
        if not is_new_format:
            raise HTTPException(status_code=400, detail="Nur deutsches CSV-Format wird unterstützt")
        
        transactions = []
        original_rows = []
        
        for idx, row in df.iterrows():
            try:
                customer_id = str(row['Kundennummer'])
                transaction_id = str(row['Unique Transaktion ID'])
                customer_name = str(row['Vollständiger Name'])
                
                amount_str = str(row['Auftragsvolumen']).replace(',', '.')
                transaction_amount = float(amount_str)
                
                art = str(row['Art']).strip()
                if art == "Kredit":
                    payment_method = "Kreditkarte"
                elif art == "Bar":
                    payment_method = "Bar"
                elif art == "SEPA":
                    payment_method = "SEPA"
                else:
                    payment_method = art
                
                in_out = str(row['In/Out']).strip()
                transaction_type = "investment" if in_out == "In" else "auszahlung"
                
                timestamp = None
                if 'Timestamp' in df.columns and pd.notna(row['Timestamp']):
                    try:
                        date_str = str(row['Timestamp'])
                        timestamp = pd.to_datetime(date_str, format='%d.%m.%Y')
                        
                        if 'Uhrzeit' in df.columns and pd.notna(row['Uhrzeit']):
                            try:
                                time_str = str(row['Uhrzeit']).replace(',', '.')
                                time_decimal = float(time_str)
                                hours = int(time_decimal * 24)
                                minutes = int((time_decimal * 24 - hours) * 60)
                                seconds = int(((time_decimal * 24 - hours) * 60 - minutes) * 60)
                                timestamp = timestamp.replace(hour=hours, minute=minutes, second=seconds)
                            except:
                                pass
                    except:
                        timestamp = None
                
                txn = Transaction(
                    customer_id=customer_id,
                    transaction_id=transaction_id,
                    customer_name=customer_name,
                    transaction_amount=transaction_amount,
                    payment_method=payment_method,
                    transaction_type=transaction_type,
                    timestamp=timestamp
                )
                transactions.append(txn)
                original_rows.append(row)
                
            except Exception as e:
                logger.warning(f"Fehler beim Parsen von Zeile {idx+2}: {e}")
                continue
        
        if not transactions:
            raise HTTPException(status_code=400, detail="Keine gültigen Transaktionen gefunden")
        
        logger.info(f"{len(transactions)} Transaktionen erfolgreich geparst")
        
        # Erstelle Analyzer und analysiere
        custom_analyzer = TransactionAnalyzer(
            alpha=0.6,
            beta=0.4,
            historical_days=365,
            use_tp_sp_system=True
        )
        
        custom_analyzer.add_transactions(transactions)
        
        # Verwende 30 Tage für aktuelle Analyse
        # Kunden ohne Transaktionen in diesem Zeitfenster bekommen Default-Profil (GREEN, Score 0)
        profiles = custom_analyzer.analyze_all_customers(recent_days=30)
        
        logger.info(f"{len(profiles)} Kunden analysiert")
        
        # Erstelle Analyse-Dictionary
        customer_analysis = {}
        for profile in profiles:
            # Flags sind direkt im CustomerRiskProfile gespeichert
            flags_str = ' | '.join(profile.flags) if profile.flags else ''
            
            customer_analysis[profile.customer_id] = {
                'Risk_Level': profile.risk_level.value,
                'Suspicion_Score': round(profile.suspicion_score, 2),  # Direkt verwenden, keine Multiplikation
                'Flags': flags_str,
                'Threshold_Avoidance_Ratio_%': round(profile.weight_analysis.threshold_avoidance_ratio * 100, 1) if profile.weight_analysis else 0.0,
                'Cumulative_Large_Amount': round(profile.weight_analysis.cumulative_large_amount, 2) if profile.weight_analysis else 0.0,
                'Temporal_Density_Weeks': round(profile.weight_analysis.temporal_density_weeks, 2) if profile.weight_analysis else 0.0,
                'Layering_Score': round(profile.statistical_analysis.layering_score, 2) if profile.statistical_analysis else 0.0,
                'Entropy_Complex': 'Ja' if profile.entropy_analysis and profile.entropy_analysis.is_complex else 'Nein',
                # Trust_Score entfernt - nicht mehr verwendet
            }
        
        # Erstelle Output-DataFrame
        output_data = []
        for idx, row in enumerate(original_rows):
            customer_id = str(row['Kundennummer'])
            analysis = customer_analysis.get(customer_id, {})
            
            output_row = row.to_dict()
            output_row['Risk_Level'] = analysis.get('Risk_Level', 'GREEN')
            output_row['Suspicion_Score'] = analysis.get('Suspicion_Score', 0.0)  # Direkt verwenden
            output_row['Flags'] = analysis.get('Flags', '')
            output_row['Threshold_Avoidance_Ratio_%'] = analysis.get('Threshold_Avoidance_Ratio_%', 0.0)
            output_row['Cumulative_Large_Amount'] = analysis.get('Cumulative_Large_Amount', 0.0)
            output_row['Temporal_Density_Weeks'] = analysis.get('Temporal_Density_Weeks', 0.0)
            output_row['Layering_Score'] = analysis.get('Layering_Score', 0.0)
            output_row['Entropy_Complex'] = analysis.get('Entropy_Complex', 'Nein')
            # Trust_Score entfernt - nicht mehr verwendet
            
            output_data.append(output_row)
        
        output_df = pd.DataFrame(output_data)
        
        # Entferne Trust_Score aus DataFrame (für CSV und Excel)
        if 'Trust_Score' in output_df.columns:
            output_df = output_df.drop(columns=['Trust_Score'])
        
        # Speichere CSV
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"Analyzed_Trades_{timestamp}.csv"
        output_path = output_dir / output_filename
        
        output_df.to_csv(output_path, index=False, encoding='utf-8-sig')
        logger.info(f"Analysierte CSV gespeichert: {output_filename}")
        
        # Erstelle auch Excel-Datei (falls openpyxl verfügbar)
        excel_filename = None
        if OPENPYXL_AVAILABLE:
            try:
                excel_filename = create_excel_file(output_df, output_dir, timestamp)
                logger.info(f"Analysierte Excel gespeichert: {excel_filename}")
            except Exception as e:
                logger.warning(f"Excel-Export fehlgeschlagen: {e}")
        
        # Zusammenfassung
        summary = {
            "green": sum(1 for p in profiles if p.risk_level == RiskLevel.GREEN),
            "yellow": sum(1 for p in profiles if p.risk_level == RiskLevel.YELLOW),
            "orange": sum(1 for p in profiles if p.risk_level == RiskLevel.ORANGE),
            "red": sum(1 for p in profiles if p.risk_level == RiskLevel.RED),
        }
        
        return {
            "status": "success",
            "message": f"{len(transactions)} Transaktionen analysiert",
            "analyzed_customers": len(profiles),
            "summary": summary,
            "csv_filename": output_filename,
            "excel_filename": excel_filename
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Fehler bei CSV-Analyse: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Fehler: {str(e)}")


def create_excel_file(df: pd.DataFrame, output_dir: Path, timestamp: str) -> str:
    """
    Erstellt eine formatierte Excel-Datei aus dem DataFrame
    
    Args:
        df: DataFrame mit Analyse-Daten
        output_dir: Ausgabe-Verzeichnis
        timestamp: Zeitstempel für Dateinamen
        
    Returns:
        Dateiname der erstellten Excel-Datei
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl nicht verfügbar")
    
    excel_filename = f"Analyzed_Trades_{timestamp}.xlsx"
    excel_path = output_dir / excel_filename
    
    # Erstelle Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Analyse-Ergebnisse"
    
    # Definiere Farben für Risk Levels
    risk_colors = {
        'GREEN': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'YELLOW': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'ORANGE': PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid'),  # Orange (#FFA500)
        'RED': PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  # Helleres Rot für bessere Lesbarkeit
    }
    
    # Header-Formatierung
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Schreibe Header
    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Finde Spalten-Indizes für spezielle Formatierung
    risk_level_col_idx = headers.index('Risk_Level') + 1 if 'Risk_Level' in headers else None
    suspicion_score_col_idx = headers.index('Suspicion_Score') + 1 if 'Suspicion_Score' in headers else None
    datum_col_idx = headers.index('Datum') + 1 if 'Datum' in headers else None
    uhrzeit_col_idx = headers.index('Uhrzeit') + 1 if 'Uhrzeit' in headers else None
    timestamp_col_idx = headers.index('Timestamp') + 1 if 'Timestamp' in headers else None
    auftragsvolumen_col_idx = headers.index('Auftragsvolumen') + 1 if 'Auftragsvolumen' in headers else None
    
    # Schreibe Daten
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            header_name = headers[col_idx - 1]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Risk Level Spalte einfärben
            if col_idx == risk_level_col_idx and value in risk_colors:
                cell.fill = risk_colors[value]
                cell.font = Font(bold=True, color='000000')  # Schwarze Schrift für bessere Lesbarkeit
            
            # Datum formatieren
            if col_idx == datum_col_idx:
                try:
                    # Versuche Datum zu parsen
                    if isinstance(value, str):
                        # Format: DD.MM.YYYY
                        from datetime import datetime as dt
                        date_obj = dt.strptime(value, '%d.%m.%Y')
                        cell.value = date_obj
                        cell.number_format = 'DD.MM.YYYY'
                    elif isinstance(value, pd.Timestamp):
                        cell.value = value
                        cell.number_format = 'DD.MM.YYYY'
                except:
                    pass  # Falls Parsing fehlschlägt, Original-Wert behalten
            
            # Uhrzeit formatieren
            elif col_idx == uhrzeit_col_idx:
                try:
                    if isinstance(value, str):
                        # Format könnte variieren, versuche verschiedene Formate
                        from datetime import datetime as dt
                        # Versuche zu parsen (kann Komma oder Punkt als Dezimaltrenner haben)
                        if ',' in str(value) or '.' in str(value):
                            # Numerischer Wert (Tagesbruchteil)
                            time_value = float(str(value).replace(',', '.'))
                            # Konvertiere zu Excel-Zeit (0.0 = 00:00:00, 0.5 = 12:00:00)
                            cell.value = time_value
                            cell.number_format = 'HH:MM:SS'
                        else:
                            # Text-Format, versuche zu parsen
                            cell.value = value
                            cell.number_format = 'HH:MM:SS'
                except:
                    pass
            
            # Timestamp formatieren (Zahlenformat)
            elif col_idx == timestamp_col_idx:
                try:
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.000000000000000'  # Viele Dezimalstellen für Timestamp
                    elif isinstance(value, str):
                        # Versuche zu float zu konvertieren
                        cell.value = float(str(value).replace(',', '.'))
                        cell.number_format = '0.000000000000000'
                except:
                    pass
            
            # Auftragsvolumen: Buchhaltungszahlenformat
            elif col_idx == auftragsvolumen_col_idx:
                try:
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'  # Buchhaltungsformat mit Tausender-Trennzeichen
                    elif isinstance(value, str):
                        # Versuche zu float zu konvertieren (Komma zu Punkt)
                        cell.value = float(str(value).replace(',', '.'))
                        cell.number_format = '#,##0.00'
                except:
                    pass
            
            # Suspicion Score: Zahlenformat, negativ
            elif col_idx == suspicion_score_col_idx:
                try:
                    if isinstance(value, (int, float)):
                        cell.value = value * -1  # Nur negativ machen
                        cell.number_format = '0.00'  # 2 Dezimalstellen
                    elif isinstance(value, str):
                        cell.value = float(str(value).replace(',', '.')) * -1
                        cell.number_format = '0.00'
                except:
                    pass
            
            # Prozentwerte formatieren
            elif '%' in header_name and isinstance(value, (int, float)):
                cell.number_format = '0.0'
            
            # Weitere Beträge formatieren
            elif header_name in ['Cumulative_Large_Amount'] and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
    
    # Spaltenbreiten anpassen
    column_widths = {
        'Datum': 12,
        'Uhrzeit': 10,
        'Timestamp': 18,
        'Kundennummer': 14,
        'Unique Transaktion ID': 25,
        'Vollständiger Name': 25,
        'Auftragsvolumen': 15,
        'In/Out': 8,
        'Art': 12,
        'Risk_Level': 12,
        'Suspicion_Score': 15,
        'Flags': 50,
        'Threshold_Avoidance_Ratio_%': 25,
        'Cumulative_Large_Amount': 20,
        'Temporal_Density_Weeks': 20,
        'Layering_Score': 15,
        'Entropy_Complex': 15
        # Trust_Score entfernt - nicht mehr verwendet
    }
    
    for col_idx, header in enumerate(headers, start=1):
        width = column_widths.get(header, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Header-Zeile einfrieren
    ws.freeze_panes = 'A2'
    
    # Speichere Excel-Datei
    wb.save(excel_path)
    
    return excel_filename


@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """
    Download einer generierten CSV- oder Excel-Datei
    """
    try:
        file_path = output_dir / filename
        
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="Datei nicht gefunden")
        
        # Sicherheitscheck: Nur Dateien im output_dir
        if not str(file_path.resolve()).startswith(str(output_dir.resolve())):
            raise HTTPException(status_code=403, detail="Zugriff verweigert")
        
        # Bestimme Media-Type
        if filename.endswith('.xlsx'):
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        else:
            media_type = 'text/csv'
        
        logger.info(f"Datei-Download: {filename}")
        return FileResponse(
            path=file_path,
            filename=filename,
            media_type=media_type
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Fehler beim Download: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    print("""
    ╔═══════════════════════════════════════════════════════════════╗
    ║                                                               ║
    ║     CLARA - Transaction Analysis System                      ║
    ║     Anti-Smurfing & Anomaly Detection                        ║
    ║                                                               ║
    ║     API läuft auf: http://localhost:8000                     ║
    ║     Dokumentation: http://localhost:8000/docs                ║
    ║                                                               ║
    ╚═══════════════════════════════════════════════════════════════╝
    """)
    
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8000,
        log_level="info"
    )

